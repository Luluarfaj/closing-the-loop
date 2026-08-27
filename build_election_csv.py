#!/usr/bin/env python3
"""
Build election_2024_by_state.csv from the FEC's official 2024 results.

WHY THIS FILE EXISTS
  The political confound is the most important empirical result in this project:
  flu concern tracks partisan lean at rho -0.688 (p = 2.4e-08), far more strongly
  than it tracks actual flu (-0.268). A result that big has to rest on real data
  from a citable source, not on numbers anyone typed from memory.

  It previously did rest on memory. Those numbers had Oklahoma at +39; the official
  two-party margin is +34.9. The confound came out identical either way, which is a
  warning rather than a comfort: a large effect hides input errors instead of
  surfacing them.

SOURCE
  Federal Election Commission, "Official 2024 Presidential General Election
  Results", compiled 2025-01-16. Public Records Branch, FEC.
  https://www.fec.gov/documents/5645/2024presgeresults.xlsx
  In-file provenance: "Popular Vote: State Elections Offices; Electoral Vote: State
  Certificates of Vote, U.S. National Archives."
  US federal government work; no copyright asserted.

  Cross-checked against the MIT Election Data and Science Lab mirror
  (github.com/MEDSL/2024-elections-official, 2024-president-state.csv). The two agree
  vote-for-vote on all 51 jurisdictions. If you would rather cite MEDSL, the
  canonical deposit is https://doi.org/10.7910/DVN/42MVDX (CC0).

THE BASIS, STATED EXPLICITLY (a reviewer will ask)
  margin = (Trump - Harris) / (Trump + Harris) * 100
  Two-party only. Third-party, write-in, and scattering votes are EXCLUDED from the
  denominator. Positive = Republican advantage. DC comes out at -86.6 on this basis
  (it is -83.8 if you use all votes cast as the denominator). Pick one basis, say
  which, and stay with it.

NOTE ON NEW YORK AND CONNECTICUT
  Both use fusion voting: the same candidate appears on multiple party lines. The FEC
  sheet reports one consolidated total per candidate, so summing its HARRIS and TRUMP
  columns is correct and drops nothing. (If you ever switch to MEDSL, aggregate on
  party_simplified, not party_detailed - NY's Harris row is labelled
  "WORKING FAMILIES / DEMOCRAT" and a party_detailed=='DEMOCRAT' filter silently
  returns zero rows for NY.)

Run:  python3 build_election_csv.py
"""
import csv
import subprocess
import sys

URL = "https://www.fec.gov/documents/5645/2024presgeresults.xlsx"
XLSX = "fec_2024_pres.xlsx"
OUT = "election_2024_by_state.csv"


def main():
    try:
        import openpyxl
    except ImportError:
        sys.exit("pip install openpyxl")

    print(f"fetching {URL}")
    r = subprocess.run(["curl", "-sS", "-L", "-o", XLSX, URL,
                        "-w", "%{http_code} %{size_download}"],
                       capture_output=True, text=True, timeout=120)
    print(f"  http {r.stdout.strip()} bytes -> {XLSX}")

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    print(f"  sheet '{wb.sheetnames[0]}' {ws.max_row} x {ws.max_column}")

    hdr = [c.value for c in ws[1]]
    try:
        iS, iH, iT = hdr.index("STATE"), hdr.index("HARRIS"), hdr.index("TRUMP")
    except ValueError:
        sys.exit(f"  unexpected header layout: {hdr[:8]}\n"
                 "  the FEC may have republished the file; check the columns.")

    rows = []
    for r_ in ws.iter_rows(min_row=2, values_only=True):
        st = r_[iS]
        # Rows 53+ are 'Total:', 'Percentage:', and footnotes. Real jurisdictions are
        # the 2-letter codes; everything else is filtered by length.
        if not st or not isinstance(st, str) or len(st.strip()) != 2:
            continue
        h = r_[iH] or 0      # blank means zero votes, not missing
        t = r_[iT] or 0
        if (h + t) == 0:
            continue
        rows.append((st.strip().lower(), 100.0 * (t - h) / (t + h)))

    if len(rows) != 51:
        print(f"  WARNING: parsed {len(rows)} jurisdictions, expected 51 (50 states + DC).")

    rows.sort()
    with open(OUT, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["state", "gop_margin", "source"])
        for s, m in rows:
            w.writerow([s, round(m, 2), "FEC 2024 official, two-party basis"])

    d = dict(rows)
    print(f"\n  parsed {len(rows)} jurisdictions. spot checks (R minus D, two-party):")
    for s in ["dc", "wy", "wv", "ok", "ny", "ct", "ca", "tx"]:
        if s in d:
            print(f"    {s.upper():3} {d[s]:+6.1f}")
    print(f"\n  saved -> {OUT}")
    print("  NEXT: python3 compute_d.py")


if __name__ == "__main__":
    main()
